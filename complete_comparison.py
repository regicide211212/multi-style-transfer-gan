import os
import numpy as np
import cv2
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
import glob
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def calculate_metrics(original_img, processed_img):
    """Calculate image quality metrics"""
    # Ensure the same dimensions
    if original_img.shape != processed_img.shape:
        processed_img = cv2.resize(processed_img, (original_img.shape[1], original_img.shape[0]))
    
    # Convert to float and normalize to 0-1 range
    original_img = original_img.astype(float) / 255.0
    processed_img = processed_img.astype(float) / 255.0
    
    # Calculate MSE
    mse_value = np.mean((original_img - processed_img) ** 2)
    
    # Calculate PSNR
    psnr_value = psnr(original_img, processed_img, data_range=1.0)
    
    # Calculate SSIM
    ssim_value = ssim(original_img, processed_img, channel_axis=2, data_range=1.0)
    
    return mse_value, psnr_value, ssim_value

def compare_with_original():
    # Folder paths
    original_folder = "E:\\desktop\\graduate\\gan_proj_0217_copy\\monet2photo_uvcgan2\\test_images"
    cyclegan_folder = "E:\\desktop\\graduate\\gan_proj_0217_copy\\monet2photo_uvcgan2\\results\\cyclegan_monet2photo"
    local_style_folder = "E:\\desktop\\graduate\\gan_proj_0217_copy\\monet2photo_uvcgan2\\results\\local_style_enhanced_monet2photo"
    
    # Get all image files
    extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp']
    
    original_images = []
    for ext in extensions:
        original_images.extend(glob.glob(os.path.join(original_folder, ext)))
    
    cyclegan_images = []
    for ext in extensions:
        cyclegan_images.extend(glob.glob(os.path.join(cyclegan_folder, ext)))
    
    local_style_images = []
    for ext in extensions:
        local_style_images.extend(glob.glob(os.path.join(local_style_folder, ext)))
    
    print(f"Original photos: Found {len(original_images)} images")
    print(f"CycleGAN processed: Found {len(cyclegan_images)} images")
    print(f"LocalStyle processed: Found {len(local_style_images)} images")
    
    # Use filename as dictionary keys
    original_dict = {os.path.basename(img): img for img in original_images}
    cyclegan_dict = {os.path.basename(img): img for img in cyclegan_images}
    local_style_dict = {os.path.basename(img): img for img in local_style_images}
    
    # Find matching images across all three folders
    common_filenames = []
    for orig_name in original_dict.keys():
        # Try exact match first
        found_in_cyclegan = False
        found_in_local_style = False
        
        for cy_name in cyclegan_dict.keys():
            if orig_name == cy_name or orig_name in cy_name:
                found_in_cyclegan = cy_name
                break
                
        for ls_name in local_style_dict.keys():
            if orig_name == ls_name or orig_name in ls_name:
                found_in_local_style = ls_name
                break
                
        if found_in_cyclegan and found_in_local_style:
            common_filenames.append((orig_name, found_in_cyclegan, found_in_local_style))
    
    print(f"Found {len(common_filenames)} comparable image sets")
    
    if len(common_filenames) == 0:
        print("No complete image matches found, please check file naming")
        return
    
    # Calculate metrics
    cyclegan_metrics = {"mse": [], "psnr": [], "ssim": []}
    local_style_metrics = {"mse": [], "psnr": [], "ssim": []}
    comparative_results = []
    
    # Create a list to store data for Excel
    excel_data = []
    
    for orig_name, cy_name, ls_name in common_filenames:
        orig_path = original_dict[orig_name]
        cy_path = cyclegan_dict[cy_name]
        ls_path = local_style_dict[ls_name]
        
        orig_img = cv2.imread(orig_path)
        cy_img = cv2.imread(cy_path)
        ls_img = cv2.imread(ls_path)
        
        if orig_img is None or cy_img is None or ls_img is None:
            print(f"Unable to read image set: {orig_name}")
            continue
        
        # CycleGAN metrics
        cy_mse, cy_psnr, cy_ssim = calculate_metrics(orig_img, cy_img)
        cyclegan_metrics["mse"].append(cy_mse)
        cyclegan_metrics["psnr"].append(cy_psnr)
        cyclegan_metrics["ssim"].append(cy_ssim)
        
        # LocalStyle metrics
        ls_mse, ls_psnr, ls_ssim = calculate_metrics(orig_img, ls_img)
        local_style_metrics["mse"].append(ls_mse)
        local_style_metrics["psnr"].append(ls_psnr)
        local_style_metrics["ssim"].append(ls_ssim)
        
        # Save comparison results
        result = {
            "orig_name": orig_name,
            "cyclegan": {
                "name": cy_name,
                "mse": cy_mse,
                "psnr": cy_psnr,
                "ssim": cy_ssim
            },
            "local_style": {
                "name": ls_name,
                "mse": ls_mse,
                "psnr": ls_psnr,
                "ssim": ls_ssim
            },
            "comparison": {
                "mse_better": "LocalStyle" if ls_mse < cy_mse else "CycleGAN",
                "psnr_better": "LocalStyle" if ls_psnr > cy_psnr else "CycleGAN",
                "ssim_better": "LocalStyle" if ls_ssim > cy_ssim else "CycleGAN"
            }
        }
        comparative_results.append(result)
        
        # Add data for Excel
        excel_row = {
            "Image Name": orig_name,
            "CycleGAN MSE": cy_mse,
            "LocalStyle MSE": ls_mse,
            "MSE Better": "LocalStyle" if ls_mse < cy_mse else "CycleGAN",
            "CycleGAN PSNR": cy_psnr,
            "LocalStyle PSNR": ls_psnr,
            "PSNR Better": "LocalStyle" if ls_psnr > cy_psnr else "CycleGAN",
            "CycleGAN SSIM": cy_ssim,
            "LocalStyle SSIM": ls_ssim,
            "SSIM Better": "LocalStyle" if ls_ssim > cy_ssim else "CycleGAN"
        }
        excel_data.append(excel_row)
        
        # Output comparison for each image set
        print(f"\nOriginal image: {orig_name}")
        print(f"CycleGAN processed ({cy_name}):")
        print(f"  MSE: {cy_mse:.6f}, PSNR: {cy_psnr:.2f}dB, SSIM: {cy_ssim:.4f}")
        print(f"LocalStyle processed ({ls_name}):")
        print(f"  MSE: {ls_mse:.6f}, PSNR: {ls_psnr:.2f}dB, SSIM: {ls_ssim:.4f}")
        print(f"Comparison results:")
        print(f"  MSE: {'LocalStyle better' if ls_mse < cy_mse else 'CycleGAN better'} (diff: {abs(ls_mse - cy_mse):.6f})")
        print(f"  PSNR: {'LocalStyle better' if ls_psnr > cy_psnr else 'CycleGAN better'} (diff: {abs(ls_psnr - cy_psnr):.2f}dB)")
        print(f"  SSIM: {'LocalStyle better' if ls_ssim > cy_ssim else 'CycleGAN better'} (diff: {abs(ls_ssim - cy_ssim):.4f})")
    
    # Calculate averages
    avg_cy_mse = np.mean(cyclegan_metrics["mse"])
    avg_cy_psnr = np.mean(cyclegan_metrics["psnr"])
    avg_cy_ssim = np.mean(cyclegan_metrics["ssim"])
    
    avg_ls_mse = np.mean(local_style_metrics["mse"])
    avg_ls_psnr = np.mean(local_style_metrics["psnr"])
    avg_ls_ssim = np.mean(local_style_metrics["ssim"])
    
    # Count advantages
    cy_wins = {"mse": 0, "psnr": 0, "ssim": 0}
    ls_wins = {"mse": 0, "psnr": 0, "ssim": 0}
    
    for result in comparative_results:
        if result["comparison"]["mse_better"] == "CycleGAN":
            cy_wins["mse"] += 1
        else:
            ls_wins["mse"] += 1
            
        if result["comparison"]["psnr_better"] == "CycleGAN":
            cy_wins["psnr"] += 1
        else:
            ls_wins["psnr"] += 1
            
        if result["comparison"]["ssim_better"] == "CycleGAN":
            cy_wins["ssim"] += 1
        else:
            ls_wins["ssim"] += 1
    
    total_images = len(comparative_results)
    
    # Print summary
    print("\n" + "=" * 60)
    print("Quality Assessment Summary:")
    print("=" * 60)
    print("Average Metrics Comparison (relative to original images):")
    print(f"  CycleGAN:")
    print(f"    MSE: {avg_cy_mse:.6f}")
    print(f"    PSNR: {avg_cy_psnr:.2f}dB")
    print(f"    SSIM: {avg_cy_ssim:.4f}")
    print(f"  LocalStyle:")
    print(f"    MSE: {avg_ls_mse:.6f}")
    print(f"    PSNR: {avg_ls_psnr:.2f}dB")
    print(f"    SSIM: {avg_ls_ssim:.4f}")
    
    print("\nAdvantage Comparison:")
    print(f"  MSE (lower is better):")
    print(f"    CycleGAN better: {cy_wins['mse']} images ({cy_wins['mse']/total_images*100:.1f}%)")
    print(f"    LocalStyle better: {ls_wins['mse']} images ({ls_wins['mse']/total_images*100:.1f}%)")
    print(f"    Average difference: {abs(avg_cy_mse - avg_ls_mse):.6f} ({abs(avg_cy_mse - avg_ls_mse)/max(avg_cy_mse, avg_ls_mse)*100:.1f}%)")
    
    print(f"  PSNR (higher is better):")
    print(f"    CycleGAN better: {cy_wins['psnr']} images ({cy_wins['psnr']/total_images*100:.1f}%)")
    print(f"    LocalStyle better: {ls_wins['psnr']} images ({ls_wins['psnr']/total_images*100:.1f}%)")
    print(f"    Average difference: {abs(avg_cy_psnr - avg_ls_psnr):.2f}dB ({abs(avg_cy_psnr - avg_ls_psnr)/max(avg_cy_psnr, avg_ls_psnr)*100:.1f}%)")
    
    print(f"  SSIM (higher is better):")
    print(f"    CycleGAN better: {cy_wins['ssim']} images ({cy_wins['ssim']/total_images*100:.1f}%)")
    print(f"    LocalStyle better: {ls_wins['ssim']} images ({ls_wins['ssim']/total_images*100:.1f}%)")
    print(f"    Average difference: {abs(avg_cy_ssim - avg_ls_ssim):.4f} ({abs(avg_cy_ssim - avg_ls_ssim)/max(avg_cy_ssim, avg_ls_ssim)*100:.1f}%)")
    
    print("\nOverall Conclusion:")
    better_method = ""
    if ls_wins["mse"] > cy_wins["mse"] and ls_wins["psnr"] > cy_wins["psnr"] and ls_wins["ssim"] > cy_wins["ssim"]:
        better_method = "LocalStyle performs better than CycleGAN on all metrics"
    elif cy_wins["mse"] > ls_wins["mse"] and cy_wins["psnr"] > ls_wins["psnr"] and cy_wins["ssim"] > ls_wins["ssim"]:
        better_method = "CycleGAN performs better than LocalStyle on all metrics"
    else:
        total_ls_wins = ls_wins["mse"] + ls_wins["psnr"] + ls_wins["ssim"]
        total_cy_wins = cy_wins["mse"] + cy_wins["psnr"] + cy_wins["ssim"]
        if total_ls_wins > total_cy_wins:
            better_method = f"LocalStyle performs better in {total_ls_wins} metric evaluations (out of {total_ls_wins + total_cy_wins})"
        elif total_cy_wins > total_ls_wins:
            better_method = f"CycleGAN performs better in {total_cy_wins} metric evaluations (out of {total_ls_wins + total_cy_wins})"
        else:
            better_method = "Both methods have their advantages, overall performance is comparable"
    
    print(f"  {better_method}")
    print("=" * 60)
    
    # Simple explanation of metrics
    print("\nMetrics Explanation:")
    print("- MSE (Mean Squared Error): Lower values indicate greater similarity to original image")
    print("- PSNR (Peak Signal-to-Noise Ratio): Higher values indicate better image quality") 
    print("- SSIM (Structural Similarity Index): Values closer to 1 indicate greater structural similarity")
    
    # Set chart style
    plt.style.use('ggplot')
    
    # Define colors
    cyclegan_color = '#3498db'  # Blue
    localstyle_color = '#e74c3c'  # Red
    
    # Draw average metrics comparison chart
    fig, axs = plt.subplots(1, 3, figsize=(15, 5))
    
    # Create labels
    labels = ['CycleGAN', 'LocalStyle']
    
    # MSE (lower is better)
    bars1 = axs[0].bar(labels, [avg_cy_mse, avg_ls_mse], color=[cyclegan_color, localstyle_color])
    axs[0].set_title('Average MSE (Lower is Better)')
    axs[0].set_ylabel('Mean Squared Error')
    
    # Add value labels
    for bar in bars1:
        height = bar.get_height()
        axs[0].text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.6f}', ha='center', va='bottom')
    
    # PSNR (higher is better)
    bars2 = axs[1].bar(labels, [avg_cy_psnr, avg_ls_psnr], color=[cyclegan_color, localstyle_color])
    axs[1].set_title('Average PSNR (Higher is Better)')
    axs[1].set_ylabel('Peak Signal-to-Noise Ratio (dB)')
    
    # Add value labels
    for bar in bars2:
        height = bar.get_height()
        axs[1].text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.2f}dB', ha='center', va='bottom')
    
    # SSIM (higher is better)
    bars3 = axs[2].bar(labels, [avg_cy_ssim, avg_ls_ssim], color=[cyclegan_color, localstyle_color])
    axs[2].set_title('Average SSIM (Higher is Better)')
    axs[2].set_ylabel('Structural Similarity')
    
    # Add value labels
    for bar in bars3:
        height = bar.get_height()
        axs[2].text(bar.get_x() + bar.get_width()/2., height,
                f'{height:.4f}', ha='center', va='bottom')
    
    plt.tight_layout()
    chart_path = "E:\\desktop\\graduate\\gan_proj_0217_copy\\monet2photo_uvcgan2\\quality_comparison_chart.png"
    plt.savefig(chart_path)
    print(f"\nComparison chart saved to: {chart_path}")
    
    # Create Excel file with detailed data
    df = pd.DataFrame(excel_data)
    
    # Add average row
    average_row = {
        "Image Name": "AVERAGE",
        "CycleGAN MSE": avg_cy_mse,
        "LocalStyle MSE": avg_ls_mse,
        "MSE Better": "LocalStyle" if avg_ls_mse < avg_cy_mse else "CycleGAN",
        "CycleGAN PSNR": avg_cy_psnr,
        "LocalStyle PSNR": avg_ls_psnr,
        "PSNR Better": "LocalStyle" if avg_ls_psnr > avg_cy_psnr else "CycleGAN",
        "CycleGAN SSIM": avg_cy_ssim,
        "LocalStyle SSIM": avg_ls_ssim,
        "SSIM Better": "LocalStyle" if avg_ls_ssim > avg_cy_ssim else "CycleGAN"
    }
    
    # Create a summary for the statistics sheet
    summary_data = [
        {"Metric": "MSE (Lower is Better)", 
         "CycleGAN Value": avg_cy_mse, 
         "LocalStyle Value": avg_ls_mse, 
         "Difference": abs(avg_cy_mse - avg_ls_mse),
         "Percentage": abs(avg_cy_mse - avg_ls_mse)/max(avg_cy_mse, avg_ls_mse)*100,
         "Better Method": "LocalStyle" if avg_ls_mse < avg_cy_mse else "CycleGAN"},
        {"Metric": "PSNR (Higher is Better)", 
         "CycleGAN Value": avg_cy_psnr, 
         "LocalStyle Value": avg_ls_psnr, 
         "Difference": abs(avg_cy_psnr - avg_ls_psnr),
         "Percentage": abs(avg_cy_psnr - avg_ls_psnr)/max(avg_cy_psnr, avg_ls_psnr)*100,
         "Better Method": "LocalStyle" if avg_ls_psnr > avg_cy_psnr else "CycleGAN"},
        {"Metric": "SSIM (Higher is Better)", 
         "CycleGAN Value": avg_cy_ssim, 
         "LocalStyle Value": avg_ls_ssim, 
         "Difference": abs(avg_cy_ssim - avg_ls_ssim),
         "Percentage": abs(avg_cy_ssim - avg_ls_ssim)/max(avg_cy_ssim, avg_ls_ssim)*100,
         "Better Method": "LocalStyle" if avg_ls_ssim > avg_cy_ssim else "CycleGAN"}
    ]
    
    summary_df = pd.DataFrame(summary_data)
    
    # Add win statistics
    win_data = [
        {"Metric": "MSE", "CycleGAN Wins": cy_wins["mse"], "LocalStyle Wins": ls_wins["mse"], 
         "CycleGAN Win %": cy_wins["mse"]/total_images*100, "LocalStyle Win %": ls_wins["mse"]/total_images*100},
        {"Metric": "PSNR", "CycleGAN Wins": cy_wins["psnr"], "LocalStyle Wins": ls_wins["psnr"], 
         "CycleGAN Win %": cy_wins["psnr"]/total_images*100, "LocalStyle Win %": ls_wins["psnr"]/total_images*100},
        {"Metric": "SSIM", "CycleGAN Wins": cy_wins["ssim"], "LocalStyle Wins": ls_wins["ssim"], 
         "CycleGAN Win %": cy_wins["ssim"]/total_images*100, "LocalStyle Win %": ls_wins["ssim"]/total_images*100}
    ]
    
    win_df = pd.DataFrame(win_data)
    
    # Save to Excel with multiple sheets
    excel_file = "E:\\desktop\\graduate\\gan_proj_0217_copy\\monet2photo_uvcgan2\\detailed_metrics_comparison.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Write detailed results
        df.to_excel(writer, sheet_name='Detailed Results', index=False)
        
        # Add the average row at the bottom
        workbook = writer.book
        worksheet = writer.sheets['Detailed Results']
        
        # Format the average row
        avg_row_idx = len(df) + 2  # +2 because Excel is 1-indexed and header row
        df_avg = pd.DataFrame([average_row])
        for r_idx, row in enumerate(dataframe_to_rows(df_avg, index=False, header=False)):
            for c_idx, value in enumerate(row):
                worksheet.cell(row=avg_row_idx, column=c_idx+1, value=value)
                cell = worksheet.cell(row=avg_row_idx, column=c_idx+1)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Write summary sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Write win statistics
        win_df.to_excel(writer, sheet_name='Win Statistics', index=False)
        
        # Format win statistics sheet
        win_sheet = writer.sheets['Win Statistics']
        for column in win_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            win_sheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Detailed metrics comparison Excel file saved to: {excel_file}")

if __name__ == "__main__":
    compare_with_original()